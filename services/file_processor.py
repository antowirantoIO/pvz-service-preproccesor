import os
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill, Border, Alignment, Font
import uuid

def convert_xls_to_xlsx(xls_file, output_file):
    app = xw.App(visible=False)
    workbook = app.books.open(xls_file)

    sheet = workbook.sheets[0]

    # Unhide all rows and columns
    sheet.api.Rows.Hidden = False
    sheet.api.Columns.Hidden = False

    # Save the file as .xlsx
    workbook.save(output_file)
    workbook.close()
    app.quit()
    print(f"Converted '{xls_file}' to '{output_file}' with images preserved.")

def copy_with_xlwings(source_file, target_file):
    app = xw.App(visible=False)
    source_wb = None  # Inisialisasi variabel dengan None
    target_wb = None  # Inisialisasi variabel dengan None
    try:
        source_wb = app.books.open(source_file)
        target_wb = xw.Book()  # Create a new workbook

        for sheet in source_wb.sheets:
            new_sheet = target_wb.sheets.add(sheet.name)
            sheet.range("A1").expand().copy(new_sheet.range("A1"))

        target_wb.save(target_file)
    finally:
        if source_wb:
            source_wb.close()
        if target_wb:
            target_wb.close()
        app.quit()
        app.kill()  # Paksa Excel untuk menutup jika masih terbuka

def unmerge_cells_but_keep_values(sheet):
    for merged_cell_range in list(sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_cell_value

def remove_duplicate_words(header):
    words = header.split()
    seen = set()
    result = []
    for word in words:
        if word not in seen:
            seen.add(word)
            result.append(word)
    return " ".join(result)

def merge_headers(sheet, header_rows):
    headers = []
    for col in range(1, sheet.max_column + 1):
        combined_header = []
        for row in header_rows:
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                combined_header.append(cell_value.strip())
        unique_header = remove_duplicate_words(" ".join(combined_header))
        headers.append(unique_header)
    return headers

def clear_styles_for_rows(sheet, rows):
    for row in rows:
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = Border()  # Clear borders
            cell.fill = PatternFill(fill_type=None)  # Clear background color
            cell.alignment = Alignment(horizontal='left', vertical='top')  # Reset alignment
            cell.font = Font()  # Reset font style

def update_header_row(sheet, header_row_num, headers):
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row_num, column=col)
        cell.value = header
        cell.border = Border()  # Clear borders
        cell.fill = PatternFill(fill_type=None)  # Clear background color
        cell.alignment = Alignment(horizontal='left', vertical='top')  # Reset alignment
        cell.font = Font()  # Reset font style

def clear_extra_header_rows(sheet, header_rows):
    for row in header_rows[:-1]:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None

def auto_adjust_column_width(sheet, columns_with_images):
    for col in sheet.columns:
        column_letter = col[0].column_letter
        if column_letter in columns_with_images:
            continue
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def set_row_heights(sheet, height=30):
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = height

def process_file(file_path, processed_folder, image_folder):
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == '.xls':
        xlsx_file = 'converted_file.xlsx'
        convert_xls_to_xlsx(file_path, xlsx_file)
    else:
        xlsx_file = 'cleaned_file.xlsx'
        copy_with_xlwings(file_path, xlsx_file)

    workbook = load_workbook(filename=xlsx_file)
    sheet = workbook.active
    unmerge_cells_but_keep_values(sheet)

    image_paths = []

    # Menemukan kolom terakhir yang benar-benar berisi data
    last_used_column = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row[0].row, column=col).value is not None:
                last_used_column = max(last_used_column, col)

    # Kolom berikutnya setelah kolom terakhir yang berisi data
    image_filename_column = last_used_column + 1
    sheet.cell(row=1, column=image_filename_column).value = "Image Filename"

    rows_with_images = []
    columns_with_images = {sheet.cell(row=1, column=image_filename_column).column_letter}

    for image in sheet._images:
        # Generate a unique filename for each image
        img_filename = f"{uuid.uuid4().hex}.png"
        img_path = os.path.join(image_folder, img_filename)

        # Access the image's binary data
        img_stream = image.ref
        img_stream.seek(0)  # Ensure you're at the start of the file-like object

        # Save the image data to disk
        with open(img_path, "wb") as f:
            f.write(img_stream.read())  # Read from BytesIO and write to the file

        # Store the image path in the list
        image_paths.append(img_path)

        # Determine the row where the image is anchored
        row = image.anchor._from.row + 1  # Adjusted for zero-based index

        # Write the image filename in the last used column + 1 of the correct row
        sheet.cell(row=row, column=image_filename_column).value = img_filename

        rows_with_images.append(row)

    if rows_with_images:
        data_start_row = min(rows_with_images)
    else:
        data_start_row = 2

    header_rows = list(range(1, data_start_row))
    headers = merge_headers(sheet, header_rows)
    header_row_num = data_start_row - 1
    update_header_row(sheet, header_row_num, headers)
    clear_styles_for_rows(sheet, header_rows)
    clear_extra_header_rows(sheet, header_rows)
    auto_adjust_column_width(sheet, columns_with_images)

    output_file = os.path.join(processed_folder, "updated_" + os.path.basename(xlsx_file))
    workbook.save(output_file)
    return output_file, image_paths
